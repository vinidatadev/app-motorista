import io
from datetime import datetime, timezone
from uuid import UUID, uuid4
from typing import Any
from fastapi import APIRouter, Depends, HTTPException, Query, Request, status, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, constr
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, distinct
from database import get_db
from models import Cliente, ClienteFoto, ClienteAlteracao, ClienteEndereco, ClienteContato, User, Notificacao
from auth import require_permission
from geocode import geocode_endereco
from limiter import limiter
from notify import criar_notificacao, enviar_notificacao
import storage

router = APIRouter(tags=["clientes"])

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5 MB
MAX_CARGA_SIZE = 5 * 1024 * 1024  # 5 MB
# Limites de carga para evitar DoS via arquivo gigante / geocoding em massa.
MAX_LINHAS_CARGA = 2000
MAX_GEOCODED_POR_REQUEST = 50

# --- Schemas ---

class FotoOut(BaseModel):
    id: UUID
    url: str
    created_at: str

    @classmethod
    def from_orm(cls, f: ClienteFoto):
        # Bucket é privado: devolve uma URL temporária assinada (expira em 1h)
        # em vez da URL pública permanente armazenada no banco.
        url = f.url
        info = storage.extract_key_from_url(f.url)
        if info:
            url = storage.presign_url(info[1], bucket=info[0])
        return cls(id=f.id, url=url, created_at=f.created_at.isoformat() if f.created_at else None)


class ContatoIn(BaseModel):
    nome: constr(strip_whitespace=True, max_length=100)
    telefone: constr(strip_whitespace=True, max_length=20) | None = None


class ContatoOut(BaseModel):
    id: UUID
    nome: str
    telefone: str | None

    @classmethod
    def from_orm(cls, ct: ClienteContato):
        return cls(id=ct.id, nome=ct.nome or "", telefone=ct.telefone)


class EnderecoIn(BaseModel):
    nome: constr(strip_whitespace=True, max_length=100) | None = None
    cep: constr(strip_whitespace=True, max_length=10) | None = None
    rua: constr(strip_whitespace=True, max_length=150) | None = None
    numero: constr(strip_whitespace=True, max_length=20) | None = None
    bairro: constr(strip_whitespace=True, max_length=100) | None = None
    cidade: constr(strip_whitespace=True, max_length=100) | None = None
    estado: constr(strip_whitespace=True, min_length=2, max_length=2) | None = None
    latitude: float | None = None
    longitude: float | None = None
    ponto_referencia: constr(strip_whitespace=True, max_length=200) | None = None
    observacao: constr(strip_whitespace=True, max_length=2000) | None = None
    contatos: list[ContatoIn] = []


class EnderecoOut(BaseModel):
    id: UUID
    nome: str | None
    ordem: int = 0
    cep: str | None
    rua: str | None
    numero: str | None
    bairro: str | None
    cidade: str | None
    estado: str | None
    latitude: float | None
    longitude: float | None
    ponto_referencia: str | None
    observacao: str | None
    contatos: list[ContatoOut] = []

    @classmethod
    def from_orm(cls, e: ClienteEndereco, contatos: list[ClienteContato] | None = None):
        return cls(
            id=e.id,
            nome=e.nome,
            ordem=e.ordem,
            cep=e.cep,
            rua=e.rua,
            numero=e.numero,
            bairro=e.bairro,
            cidade=e.cidade,
            estado=e.estado,
            latitude=float(e.latitude) if e.latitude is not None else None,
            longitude=float(e.longitude) if e.longitude is not None else None,
            ponto_referencia=e.ponto_referencia,
            observacao=e.observacao,
            contatos=[ContatoOut.from_orm(ct) for ct in (contatos or [])],
        )


class ClienteOut(BaseModel):
    id: UUID
    codigo: str | None
    nome_razao_social: str
    telefone: str | None
    pessoa_contato: str | None
    cep: str | None
    rua: str | None
    numero: str | None
    bairro: str | None
    cidade: str | None
    estado: str | None
    latitude: float | None
    longitude: float | None
    ponto_referencia: str | None
    observacao: str | None
    status_endereco: str = "aprovado"
    alterado_por_nome: str | None = None
    alterado_por_empresa: str | None = None
    alterado_em: str | None = None
    fotos: list[FotoOut] = []
    enderecos: list[EnderecoOut] = []
    updated_at: str

    @classmethod
    def from_orm(
        cls,
        c: Cliente,
        fotos: list[ClienteFoto] | None = None,
        enderecos: list[ClienteEndereco] | None = None,
        contatos_por_endereco: dict | None = None,
    ):
        contatos_por_endereco = contatos_por_endereco or {}
        return cls(
            id=c.id,
            codigo=c.codigo,
            nome_razao_social=c.nome_razao_social,
            telefone=c.telefone,
            pessoa_contato=c.pessoa_contato,
            cep=c.cep,
            rua=c.rua,
            numero=c.numero,
            bairro=c.bairro,
            cidade=c.cidade,
            estado=c.estado,
            latitude=float(c.latitude) if c.latitude is not None else None,
            longitude=float(c.longitude) if c.longitude is not None else None,
            ponto_referencia=c.ponto_referencia,
            observacao=c.observacao,
            status_endereco=c.status_endereco or "aprovado",
            alterado_por_nome=c.alterado_por_nome,
            alterado_por_empresa=c.alterado_por_empresa,
            alterado_em=c.alterado_em.isoformat() if c.alterado_em else None,
            fotos=[FotoOut.from_orm(f) for f in (fotos or [])],
            enderecos=[
                EnderecoOut.from_orm(
                    e, contatos=contatos_por_endereco.get(e.id, [])
                )
                for e in (enderecos or [])
            ],
            updated_at=c.updated_at.isoformat() if c.updated_at else None,
        )


async def _carregar_fotos(db: AsyncSession, cliente_id: UUID) -> list[ClienteFoto]:
    """Carrega fotos ordenadas por data de criacao (mais recente primeiro)."""
    result = await db.execute(
        select(ClienteFoto)
        .where(ClienteFoto.cliente_id == cliente_id)
        .order_by(ClienteFoto.created_at.desc())
    )
    return result.scalars().all()


async def _carregar_enderecos(db: AsyncSession, cliente_id: UUID) -> list[ClienteEndereco]:
    """Carrega enderecos de um cliente (ordenados por ordem/ criacao)."""
    result = await db.execute(
        select(ClienteEndereco)
        .where(ClienteEndereco.cliente_id == cliente_id)
        .order_by(ClienteEndereco.ordem, ClienteEndereco.created_at)
    )
    return result.scalars().all()


async def _carregar_contatos(db: AsyncSession, endereco_ids: list[UUID]) -> dict[UUID, list[ClienteContato]]:
    """Carrega contatos de varios enderecos de uma vez: {endereco_id: [contato,...]}."""
    if not endereco_ids:
        return {}
    result = await db.execute(
        select(ClienteContato)
        .where(ClienteContato.endereco_id.in_(endereco_ids))
        .order_by(ClienteContato.created_at)
    )
    por_endereco: dict[UUID, list[ClienteContato]] = {}
    for ct in result.scalars().all():
        por_endereco.setdefault(ct.endereco_id, []).append(ct)
    return por_endereco


async def _cliente_out(db: AsyncSession, c: Cliente) -> ClienteOut:
    """Monta ClienteOut com fotos e enderecos/contatos anexos."""
    fotos = await _carregar_fotos(db, c.id)
    enderecos = await _carregar_enderecos(db, c.id)
    contatos = await _carregar_contatos(db, [e.id for e in enderecos])
    return ClienteOut.from_orm(c, fotos=fotos, enderecos=enderecos, contatos_por_endereco=contatos)


class ClienteCreate(BaseModel):
    codigo: constr(strip_whitespace=True, max_length=50) | None = None
    nome_razao_social: str = Field(..., min_length=1, max_length=150)
    telefone: constr(strip_whitespace=True, max_length=20) | None = None
    pessoa_contato: constr(strip_whitespace=True, max_length=100) | None = None
    cep: constr(strip_whitespace=True, max_length=10) | None = None
    rua: constr(strip_whitespace=True, max_length=150) | None = None
    numero: constr(strip_whitespace=True, max_length=20) | None = None
    bairro: constr(strip_whitespace=True, max_length=100) | None = None
    cidade: constr(strip_whitespace=True, max_length=100) | None = None
    estado: constr(strip_whitespace=True, min_length=2, max_length=2) | None = None
    latitude: float | None = None
    longitude: float | None = None
    ponto_referencia: constr(strip_whitespace=True, max_length=200) | None = None
    observacao: constr(strip_whitespace=True, max_length=2000) | None = None
    # Novos enderecos (listas de EnderecoIn). O primeiro vira o principal.
    enderecos: list[EnderecoIn] = []


class ClienteUpdate(BaseModel):
    codigo: constr(strip_whitespace=True, max_length=50) | None = None
    nome_razao_social: str | None = Field(default=None, min_length=1, max_length=150)
    telefone: constr(strip_whitespace=True, max_length=20) | None = None
    pessoa_contato: constr(strip_whitespace=True, max_length=100) | None = None
    cep: constr(strip_whitespace=True, max_length=10) | None = None
    rua: constr(strip_whitespace=True, max_length=150) | None = None
    numero: constr(strip_whitespace=True, max_length=20) | None = None
    bairro: constr(strip_whitespace=True, max_length=100) | None = None
    cidade: constr(strip_whitespace=True, max_length=100) | None = None
    estado: constr(strip_whitespace=True, min_length=2, max_length=2) | None = None
    latitude: float | None = None
    longitude: float | None = None
    ponto_referencia: constr(strip_whitespace=True, max_length=200) | None = None
    observacao: constr(strip_whitespace=True, max_length=2000) | None = None
    # Se enviado, substitui TODOS os enderecos/contatos do cliente.
    enderecos: list[EnderecoIn] | None = None


# --- Locais (estados / cidades) ---

@router.get("/locais/estados", response_model=list[str])
async def listar_estados(
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("visualizar")),
):
    result = await db.execute(
        select(distinct(Cliente.estado))
        .where(Cliente.estado.is_not(None))
        .order_by(Cliente.estado)
    )
    return [r[0] for r in result.all()]


@router.get("/locais/cidades", response_model=list[str])
async def listar_cidades(
    estado: str = Query(..., min_length=2, max_length=2, description="Sigla da UF"),
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("visualizar")),
):
    result = await db.execute(
        select(distinct(Cliente.cidade))
        .where(Cliente.estado == estado.upper(), Cliente.cidade.is_not(None))
        .order_by(Cliente.cidade)
    )
    return [r[0] for r in result.all()]


# --- Clientes ---

@router.get("/clientes", response_model=list[ClienteOut])
async def listar_clientes(
    estado: str | None = Query(default=None, max_length=2),
    cidade: str | None = Query(default=None, max_length=100),
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("visualizar")),
):
    stmt = select(Cliente).order_by(Cliente.estado, Cliente.cidade, Cliente.nome_razao_social)
    if estado:
        stmt = stmt.where(Cliente.estado == estado.upper())
    if cidade:
        stmt = stmt.where(Cliente.cidade.ilike(cidade))
    result = await db.execute(stmt)
    clientes = result.scalars().all()

    # Carrega fotos, enderecos e contatos em lote (evita N+1)
    ids = [c.id for c in clientes]
    fotos_por_cliente: dict[UUID, list[ClienteFoto]] = {}
    if ids:
        rf = await db.execute(
            select(ClienteFoto)
            .where(ClienteFoto.cliente_id.in_(ids))
            .order_by(ClienteFoto.created_at.desc())
        )
        for f in rf.scalars().all():
            fotos_por_cliente.setdefault(f.cliente_id, []).append(f)

    enderecos_por_cliente: dict[UUID, list[ClienteEndereco]] = {}
    if ids:
        re = await db.execute(
            select(ClienteEndereco)
            .where(ClienteEndereco.cliente_id.in_(ids))
            .order_by(ClienteEndereco.ordem, ClienteEndereco.created_at)
        )
        for e in re.scalars().all():
            enderecos_por_cliente.setdefault(e.cliente_id, []).append(e)

    todos_enderecos = [e for lista in enderecos_por_cliente.values() for e in lista]
    contatos_por_endereco = await _carregar_contatos(db, [e.id for e in todos_enderecos])

    return [
        ClienteOut.from_orm(
            c,
            fotos=fotos_por_cliente.get(c.id, []),
            enderecos=enderecos_por_cliente.get(c.id, []),
            contatos_por_endereco=contatos_por_endereco,
        )
        for c in clientes
    ]


# --- Exportar Excel ---

# Colunas de contato repetidas na planilha (nome + telefone por contato)
MAX_CONTATOS_XLSX = 3
COLUNAS_CONTATOS = [
    f"contato{i}_{suf}"
    for i in range(1, MAX_CONTATOS_XLSX + 1)
    for suf in ("nome", "telefone")
]

COLUNAS_CLIENTE_XLSX = ["codigo", "nome_razao_social", "telefone", "pessoa_contato"]
COLUNAS_ENDERECO_XLSX = ["cep", "rua", "numero", "bairro", "cidade", "estado",
                         "latitude", "longitude", "ponto_referencia", "observacao"]

COLUNAS_XLSX = (
    COLUNAS_CLIENTE_XLSX
    + ["endereco_apelido"]
    + COLUNAS_ENDERECO_XLSX
    + COLUNAS_CONTATOS
)


def _contatos_para_colunas(contatos: list) -> list:
    """Converte lista de contatos em valores para as colunas contato1..3."""
    vals = []
    for i in range(1, MAX_CONTATOS_XLSX + 1):
        ct = contatos[i - 1] if i - 1 < len(contatos) else None
        vals.append(ct.nome if ct else None)
        vals.append(ct.telefone if ct else None)
    return vals


@router.get("/clientes/export")
async def exportar_clientes(
    estado: str | None = Query(default=None, max_length=2),
    cidade: str | None = Query(default=None, max_length=100),
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("exportar")),
):
    import openpyxl

    stmt = select(Cliente).order_by(Cliente.estado, Cliente.cidade, Cliente.nome_razao_social)
    if estado:
        stmt = stmt.where(Cliente.estado == estado.upper())
    if cidade:
        stmt = stmt.where(Cliente.cidade.ilike(cidade))
    result = await db.execute(stmt)
    clientes = result.scalars().all()

    # Carrega enderecos/contatos de todos os clientes em lote
    ids = [c.id for c in clientes]
    enderecos_por_cliente: dict[UUID, list[ClienteEndereco]] = {}
    if ids:
        re = await db.execute(
            select(ClienteEndereco)
            .where(ClienteEndereco.cliente_id.in_(ids))
            .order_by(ClienteEndereco.ordem, ClienteEndereco.created_at)
        )
        for e in re.scalars().all():
            enderecos_por_cliente.setdefault(e.cliente_id, []).append(e)
    todos_enderecos = [e for lista in enderecos_por_cliente.values() for e in lista]
    contatos_por_endereco = await _carregar_contatos(db, [e.id for e in todos_enderecos])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(COLUNAS_XLSX)
    for c in clientes:
        enderecos = enderecos_por_cliente.get(c.id, [])
        if not enderecos:
            # Cliente sem endereco: uma linha so com os dados do cliente
            ws.append([
                c.codigo, c.nome_razao_social, c.telefone, c.pessoa_contato,
                None, None, None, None, None, None, None,
                None, None, None, None,
                *_contatos_para_colunas([]),
            ])
            continue
        for e in enderecos:
            contatos = contatos_por_endereco.get(e.id, [])
            ws.append([
                c.codigo, c.nome_razao_social, c.telefone, c.pessoa_contato,
                e.nome or ("Endereço principal" if e.ordem == 0 else f"Endereço {e.ordem + 1}"),
                e.cep, e.rua, e.numero, e.bairro, e.cidade, e.estado,
                float(e.latitude) if e.latitude is not None else None,
                float(e.longitude) if e.longitude is not None else None,
                e.ponto_referencia, e.observacao,
                *_contatos_para_colunas(contatos),
            ])
    # Largura auto
    for i, col in enumerate(COLUNAS_XLSX, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(len(col) + 4, 16)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clientes.xlsx"}
    )


# --- Carga em massa via Excel (preview upsert) ---
#
# Novo formato: UMA LINHA POR ENDERECO. O mesmo "codigo" pode aparecer em
# varias linhas = varios enderecos do mesmo cliente. As colunas de contato
# (contato1_nome...contato3_telefone) preenchem os contatos de cada endereco.
# As constantes de colunas (COLUNAS_XLSX, MAX_CONTATOS_XLSX, etc.) estao
# definidas na secao de export acima.

COLUNAS_NUMERICAS = {"latitude", "longitude"}


def _parse_valor(v):
    """Normaliza valor da celula: None/vazio -> None, numero -> str, floats de lat/lng ficam float."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def _parse_registro(row, idx):
    """Normaliza linha da planilha: colunas de texto viram str, lat/lng viram float.

    Evita TypeError/DataError no banco quando a celula vem como numero
    (ex.: rua/numero/codigo preenchidos como inteiro no Excel).
    """
    registro = {}
    for col, i in idx.items():
        v = _parse_valor(row[i] if i < len(row) else None)
        if v is None:
            registro[col] = None
        elif col in COLUNAS_NUMERICAS:
            try:
                registro[col] = float(v)
            except (TypeError, ValueError):
                registro[col] = None
        else:
            registro[col] = str(v)
    return registro


def _agrupar_linhas(dados_linhas: list[dict]) -> list[tuple[str, list[dict]]]:
    """Agrupa linhas por 'codigo': {codigo: [ {cliente, endereco, contatos}, ... ]}.

    Retorna lista ordenada por primeira aparicao. Cada linha vira um endereco
    do cliente (um cliente pode ter varias linhas = varios enderecos).
    """
    grupos: dict[str, list[dict]] = {}
    ordem: list[str] = []
    for d in dados_linhas:
        cod = (d.get("codigo") or "").strip()
        if not cod:
            raise HTTPException(
                status_code=422,
                detail="Toda linha precisa da coluna 'codigo' (identifica o cliente dono do endereço).",
            )
        if cod not in grupos:
            grupos[cod] = []
            ordem.append(cod)
        endereco = {k: d.get(k) for k in COLUNAS_ENDERECO_XLSX}
        endereco["nome"] = d.get("endereco_apelido")
        contatos = []
        for i in range(1, MAX_CONTATOS_XLSX + 1):
            nome = d.get(f"contato{i}_nome")
            if nome and str(nome).strip():
                contatos.append({
                    "nome": str(nome).strip(),
                    "telefone": d.get(f"contato{i}_telefone"),
                })
        grupos[cod].append({
            "cliente": {k: d.get(k) for k in COLUNAS_CLIENTE_XLSX},
            "endereco": endereco,
            "contatos": contatos,
        })
    return [(cod, grupos[cod]) for cod in ordem]


def _endereco_vazio(endereco: dict, contatos: list) -> bool:
    """Endereco vazio = nenhum campo de endereco preenchido e sem contatos."""
    if contatos:
        return False
    return not any(
        endereco.get(k)
        for k in ["cep", "rua", "numero", "bairro", "cidade", "estado",
                  "ponto_referencia", "observacao"]
    )


def _resumo_enderecos(linhas: list[dict]) -> list[dict]:
    """Resumo legivel dos enderecos de um grupo (para preview)."""
    resumo = []
    for l in linhas:
        e = l["endereco"]
        if _endereco_vazio(e, l["contatos"]):
            continue
        resumo.append({
            "nome": e.get("nome") or "Endereço principal",
            "rua": e.get("rua"),
            "numero": e.get("numero"),
            "cidade": e.get("cidade"),
            "estado": e.get("estado"),
            "contatos": l["contatos"],
        })
    return resumo


@router.post("/clientes/carga/preview")
@limiter.limit("10/minute")
async def preview_carga(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("carga")),
):
    """Faz upload do xlsx e retorna preview do que sera inserido (novos) e alterado (updates)."""
    import openpyxl

    data = await file.read()
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Envie um arquivo .xlsx")
    if len(data) > MAX_CARGA_SIZE:
        raise HTTPException(status_code=422, detail="Arquivo muito grande. Maximo 5 MB.")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Planilha invalida ou corrompida")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="Planilha vazia")
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    idx = {}
    for i, h in enumerate(header):
        if h in COLUNAS_XLSX:
            idx[h] = i

    if "nome_razao_social" not in idx:
        raise HTTPException(status_code=422, detail="Coluna 'nome_razao_social' nao encontrada")
    if "codigo" not in idx:
        raise HTTPException(status_code=422, detail="Coluna 'codigo' nao encontrada")

    dados_linhas = []
    for row in rows[1:]:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        dados_linhas.append(_parse_registro(row, idx))

    grupos = _agrupar_linhas(dados_linhas)
    codigos = [cod for cod, _ in grupos]

    existing_map = {}
    if codigos:
        result = await db.execute(select(Cliente).where(Cliente.codigo.in_(codigos)))
        for c in result.scalars().all():
            existing_map[c.codigo] = c

    novos = []
    alterados = []
    for cod, linhas in grupos:
        primeiro = linhas[0]["cliente"]
        resumo = _resumo_enderecos(linhas)

        if cod in existing_map:
            c = existing_map[cod]
            mudancas = {}
            # Compara campos de nivel cliente
            for k in ["nome_razao_social", "telefone", "pessoa_contato"]:
                v = primeiro.get(k)
                atual_val = getattr(c, k, None)
                if str(v) != str(atual_val):
                    mudancas[k] = {"de": atual_val, "para": v}
            # Compara a quantidade/enderecos
            enderecos_atual = await _carregar_enderecos(db, c.id)
            atual_resumo = [{"rua": e.rua, "numero": e.numero} for e in enderecos_atual]
            proposto_resumo = [
                {"rua": e["rua"], "numero": e["numero"]} for e in resumo
            ]
            if atual_resumo != proposto_resumo:
                mudancas["enderecos"] = {
                    "de": f"{len(enderecos_atual)} endereço(s)",
                    "para": f"{len(resumo)} endereço(s)",
                }
            alterados.append({
                "codigo": cod,
                "nome_razao_social": primeiro.get("nome_razao_social"),
                "mudancas": mudancas,
                "enderecos": resumo,
                "id": str(c.id),
            })
        else:
            novos.append({
                "codigo": cod,
                "nome_razao_social": primeiro.get("nome_razao_social"),
                "telefone": primeiro.get("telefone"),
                "cidade": resumo[0]["cidade"] if resumo else None,
                "estado": resumo[0]["estado"] if resumo else None,
                "enderecos": resumo,
            })

    return {
        "total_linhas": len(dados_linhas),
        "novos": novos,
        "alterados": alterados,
        "quantidade_novos": len(novos),
        "quantidade_alterados": len(alterados),
    }


@router.post("/clientes/carga/aplicar")
@limiter.limit("5/minute")
async def aplicar_carga(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("carga")),
):
    """Aplica a carga do xlsx (upsert por codigo; enderecos sao substituidos)."""
    import openpyxl

    data = await file.read()
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Envie um arquivo .xlsx")
    if len(data) > MAX_CARGA_SIZE:
        raise HTTPException(status_code=422, detail="Arquivo muito grande. Maximo 5 MB.")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Planilha invalida ou corrompida")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="Planilha vazia")
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx = {}
    for i, h in enumerate(header):
        if h in COLUNAS_XLSX:
            idx[h] = i
    if "nome_razao_social" not in idx:
        raise HTTPException(status_code=422, detail="Coluna 'nome_razao_social' nao encontrada")
    if "codigo" not in idx:
        raise HTTPException(status_code=422, detail="Coluna 'codigo' nao encontrada")

    dados_linhas = []
    for row in rows[1:]:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        dados_linhas.append(_parse_registro(row, idx))

    if len(dados_linhas) > MAX_LINHAS_CARGA:
        raise HTTPException(
            status_code=422,
            detail=f"Planilha com muitas linhas. Maximo permitido: {MAX_LINHAS_CARGA}."
        )

    grupos = _agrupar_linhas(dados_linhas)
    codigos = [cod for cod, _ in grupos]
    existing_map = {}
    if codigos:
        result = await db.execute(select(Cliente).where(Cliente.codigo.in_(codigos)))
        for c in result.scalars().all():
            existing_map[c.codigo] = c

    import asyncio
    inseridos = 0
    atualizados = 0
    geocoded = 0
    # Limita o nº de geocodificações por request para não segurar o worker
    # nem estourar o rate limit da API de geocoding gratuita.
    geocode_restantes = MAX_GEOCODED_POR_REQUEST
    for cod, linhas in grupos:
        # Monta lista de enderecos do grupo (com geocoding automatico)
        enderecos_data = []
        for l in linhas:
            e = dict(l["endereco"])
            contatos = list(l["contatos"])
            if _endereco_vazio(e, contatos):
                continue
            precisa_geo = (not e.get("latitude") and not e.get("longitude")) and (
                e.get("rua") or e.get("cidade") or e.get("cep")
            )
            if precisa_geo and geocode_restantes > 0:
                geocode_restantes -= 1
                coords = await geocode_endereco(
                    rua=e.get("rua"), numero=e.get("numero"), bairro=e.get("bairro"),
                    cidade=e.get("cidade"), estado=e.get("estado"), cep=e.get("cep")
                )
                if coords:
                    e["latitude"] = coords[0]
                    e["longitude"] = coords[1]
                    geocoded += 1
                # Respeita o rate limit do Nominatim (~1 req/s)
                await asyncio.sleep(1)
            enderecos_data.append({**e, "contatos": contatos})

        primeiro = linhas[0]["cliente"]
        if cod in existing_map:
            c = existing_map[cod]
            for k in ["nome_razao_social", "telefone", "pessoa_contato"]:
                v = primeiro.get(k)
                if v is not None:
                    setattr(c, k, v)
            # Carga e' operacao de admin: aplica direto (sem pendencia de aprovacao)
            c.status_endereco = "aprovado"
            c.alterado_por_user_id = None
            c.alterado_por_nome = None
            c.alterado_por_empresa = None
            c.alterado_em = None
            await _substituir_enderecos(db, c.id, enderecos_data)
            _espelhar_primeiro_endereco(c, enderecos_data)
            atualizados += 1
        else:
            if not (primeiro.get("nome_razao_social") or "").strip():
                raise HTTPException(
                    status_code=422,
                    detail=f"Código '{cod}' sem 'nome_razao_social' preenchido.",
                )
            c = Cliente(
                **{k: primeiro.get(k) for k in COLUNAS_CLIENTE_XLSX},
                status_endereco="aprovado",
            )
            db.add(c)
            await db.flush()
            _espelhar_primeiro_endereco(c, enderecos_data)
            await _substituir_enderecos(db, c.id, enderecos_data)
            inseridos += 1

    await db.commit()
    return {
        "inseridos": inseridos,
        "atualizados": atualizados,
        "geocoded": geocoded,
        "total": inseridos + atualizados
    }


# --- CRUD por ID (depois das rotas estaticas para nao conflitar) ---

CAMPOS_ENDERECO_SET = set(COLUNAS_ENDERECO_XLSX)


def _endereco_vazio_flat(d: dict) -> bool:
    """True se o dict de endereco nao tem nenhum campo de endereco preenchido."""
    return not any(d.get(k) for k in CAMPOS_ENDERECO_SET)


async def _geocode_endereco_dict(d: dict) -> None:
    """Faz forward geocoding no endereco (dict) se faltar lat/lng e houver endereco."""
    if (d.get("latitude") is None and d.get("longitude") is None) and (
        d.get("rua") or d.get("cidade") or d.get("cep")
    ):
        coords = await geocode_endereco(
            rua=d.get("rua"), numero=d.get("numero"), bairro=d.get("bairro"),
            cidade=d.get("cidade"), estado=d.get("estado"), cep=d.get("cep")
        )
        if coords:
            d["latitude"] = coords[0]
            d["longitude"] = coords[1]


def _espelhar_primeiro_endereco(c: Cliente, enderecos_data: list[dict]) -> None:
    """Espelha o primeiro endereco nos campos flat de Cliente (compatibilidade)."""
    if not enderecos_data:
        return
    primeiro = enderecos_data[0]
    for campo in CAMPOS_ENDERECO_SET:
        setattr(c, campo, primeiro.get(campo))


async def _substituir_enderecos(
    db: AsyncSession,
    cliente_id: UUID,
    enderecos_data: list[dict],
) -> None:
    """Substitui TODOS os enderecos/contatos de um cliente pelos dados enviados.

    O `enderecos_data` eh uma lista de dicts contendo os campos de endereco
    mais a chave `contatos` (lista de {nome, telefone}).
    """
    from sqlalchemy import delete as sa_delete

    # Remove enderecos atuais (e seus contatos via cascade manual)
    atual = await db.execute(
        select(ClienteEndereco).where(ClienteEndereco.cliente_id == cliente_id)
    )
    enderecos_atuais = atual.scalars().all()
    ids_atuais = [e.id for e in enderecos_atuais]
    if ids_atuais:
        await db.execute(
            sa_delete(ClienteContato).where(ClienteContato.endereco_id.in_(ids_atuais))
        )
        for e in enderecos_atuais:
            await db.delete(e)

    for i, ed in enumerate(enderecos_data):
        contatos = [dict(ct) for ct in (ed.pop("contatos", None) or [])]
        if _endereco_vazio_flat(ed) and not contatos:
            continue
        e = ClienteEndereco(
            cliente_id=cliente_id,
            ordem=i,
            **{k: ed.get(k) for k in CAMPOS_ENDERECO_SET if k in ed},
            nome=ed.get("nome"),
        )
        db.add(e)
        await db.flush()
        for ct in contatos:
            nome = (ct.get("nome") or "").strip()
            if not nome:
                continue
            db.add(ClienteContato(
                endereco_id=e.id,
                nome=nome,
                telefone=ct.get("telefone") or None,
            ))


@router.post("/clientes", response_model=ClienteOut, status_code=status.HTTP_201_CREATED)
@limiter.limit("30/minute")
async def criar_cliente(
    request: Request,
    body: ClienteCreate,
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("criar")),
):
    if body.codigo:
        existing = await db.execute(select(Cliente).where(Cliente.codigo == body.codigo))
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail=f"Cliente com codigo '{body.codigo}' ja existe")

    dados = body.model_dump()
    enderecos_data = dados.pop("enderecos", None) or []

    # Geocoding automatico para cada endereco sem coords
    for ed in enderecos_data:
        await _geocode_endereco_dict(ed)

    c = Cliente(**dados)
    db.add(c)
    await db.flush()
    _espelhar_primeiro_endereco(c, enderecos_data)
    await _substituir_enderecos(db, c.id, enderecos_data)
    await db.commit()
    await db.refresh(c)
    return await _cliente_out(db, c)


@router.get("/clientes/{cliente_id}", response_model=ClienteOut)
async def obter_cliente(
    cliente_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("visualizar")),
):
    c = await db.get(Cliente, cliente_id)
    if not c:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cliente nao encontrado")
    return await _cliente_out(db, c)


@router.put("/clientes/{cliente_id}", response_model=ClienteOut)
@limiter.limit("60/minute")
async def atualizar_cliente(
    request: Request,
    cliente_id: UUID,
    body: ClienteUpdate,
    db: AsyncSession = Depends(get_db),
    user: Any = Depends(require_permission("editar")),
):
    """Atualiza o cliente OU cria submissao pendente (se motorista sem 'aprovar').
    Quem tem 'aprovar' edita direto. Quem so' tem 'editar' gera alteracao pendente."""
    c = await db.get(Cliente, cliente_id)
    if not c:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cliente nao encontrado")

    # Bloqueia edicao enquanto houver alteracao pendente de aprovacao.
    # Depois que o aprovador aprovar/recusar, o status volta para 'aprovado' e a edicao libera.
    if c.status_endereco == "atualizando":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cliente com alteração pendente de aprovação. A edição só será liberada após a aprovação."
        )

    dados = body.model_dump(exclude_unset=True)
    enderecos_data = dados.pop("enderecos", None)

    permission_aprovar = (user["role"] == "admin") or ("aprovar" in (user.get("permissions") or []))
    alvos_aprovadores: list[tuple[User, Notificacao]] = []

    if permission_aprovar:
        # Aprovador/admin: edita direto no cliente
        for campo, valor in dados.items():
            setattr(c, campo, valor)

        if enderecos_data is not None:
            # Geocoding automatico para cada endereco sem coords
            for ed in enderecos_data:
                await _geocode_endereco_dict(ed)
            await _substituir_enderecos(db, c.id, enderecos_data)
            _espelhar_primeiro_endereco(c, enderecos_data)

        # Forward geocoding automatico (endereco flat, quando nao enviou enderecos)
        if (enderecos_data is None and "latitude" not in dados and "longitude" not in dados and
                c.latitude is None and c.longitude is None and
                (c.rua or c.cidade or c.cep)):
            coords = await geocode_endereco(
                rua=c.rua, numero=c.numero, bairro=c.bairro,
                cidade=c.cidade, estado=c.estado, cep=c.cep
            )
            if coords:
                c.latitude = coords[0]
                c.longitude = coords[1]

        # Marca como aprovado (foi alterado direto) e registra quem alterou
        c.status_endereco = "aprovado"
        c.alterado_por_user_id = UUID(user["user_id"])
        c.alterado_por_nome = user["name"]
        c.alterado_por_empresa = user.get("empresa") or "AC"
        c.alterado_em = datetime.now(timezone.utc)
    else:
        # Motorista: grava snapshot dos campos propostos e marca pendencia
        # Snapshot com TODOS os campos editaveis no estado proposto (mesclando com os atuais)
        campos_editaveis = [
            "codigo", "nome_razao_social", "telefone", "pessoa_contato", "cep", "rua",
            "numero", "bairro", "cidade", "estado", "latitude", "longitude",
            "ponto_referencia", "observacao"
        ]
        snapshot = {}
        for campo in campos_editaveis:
            if campo in dados:
                snapshot[campo] = dados[campo]
            else:
                # Manter valor atual
                v = getattr(c, campo)
                if campo in ("latitude", "longitude") and v is not None:
                    snapshot[campo] = float(v)
                else:
                    snapshot[campo] = v

        # Conjunto de enderecos propostos (o motorista sempre envia a lista completa)
        if enderecos_data is not None:
            snapshot["enderecos"] = enderecos_data
        else:
            # Mantem os enderecos atuais no snapshot (estado atual)
            enderecos_atuais = await _carregar_enderecos(db, c.id)
            snapshot["enderecos"] = [
                _endereco_para_dict(e, (await _carregar_contatos(db, [e.id])).get(e.id, []))
                for e in enderecos_atuais
            ]

        # IMPORTANTE: NAO altera os dados do cliente (snapshot fica guardado para revisao)

        # Atualiza o controle de quem submeteu
        c.status_endereco = "atualizando"
        c.alterado_por_user_id = UUID(user["user_id"])
        c.alterado_por_nome = user["name"]
        c.alterado_por_empresa = user.get("empresa") or "AC"
        c.alterado_em = datetime.now(timezone.utc)

        # Remove submissoes pendentes anteriores deste cliente (mantem apenas a mais nova)
        anteriores = await db.execute(
            select(ClienteAlteracao).where(
                ClienteAlteracao.cliente_id == cliente_id,
                ClienteAlteracao.status == "pendente"
            )
        )
        for a in anteriores.scalars().all():
            await db.delete(a)

        # Cria nova submissao pendente
        alt = ClienteAlteracao(
            cliente_id=cliente_id,
            snapshot=snapshot,
            motorista_user_id=UUID(user["user_id"]),
            motorista_nome=user["name"],
            motorista_empresa=user.get("empresa") or "AC",
            status="pendente",
        )
        db.add(alt)
        await db.flush()

        # Notifica os aprovadores (admin + aprovadores da mesma empresa)
        alvos_aprovadores = await _notificar_aprovadores(
            db,
            cliente=c,
            alteracao_id=alt.id,
            motorista_nome=user["name"],
        )

    await db.commit()
    await db.refresh(c)
    # Push em tempo real para os aprovadores notificados
    for u, n in alvos_aprovadores:
        await enviar_notificacao(u.id, n)
    return await _cliente_out(db, c)


@router.delete("/clientes/{cliente_id}", status_code=status.HTTP_204_NO_CONTENT)
async def deletar_cliente(
    cliente_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("deletar")),
):
    c = await db.get(Cliente, cliente_id)
    if not c:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cliente nao encontrado")
    # Remove fotos do MinIO antes de apagar o cliente
    fotos = await _carregar_fotos(db, cliente_id)
    for f in fotos:
        info = storage.extract_key_from_url(f.url)
        if info:
            storage.delete_file(info[1], bucket=info[0])
    # Remove registros de foto do banco
    for f in fotos:
        await db.delete(f)
    # Remove enderecos e contatos do cliente
    enderecos = await _carregar_enderecos(db, cliente_id)
    if enderecos:
        from sqlalchemy import delete as sa_delete
        await db.execute(
            sa_delete(ClienteContato).where(
                ClienteContato.endereco_id.in_([e.id for e in enderecos])
            )
        )
        for e in enderecos:
            await db.delete(e)
    await db.delete(c)
    await db.commit()


# --- Fotos do cliente ---

@router.post("/clientes/{cliente_id}/fotos", response_model=FotoOut, status_code=status.HTTP_201_CREATED)
@limiter.limit("20/minute")
async def upload_foto(
    request: Request,
    cliente_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("editar")),
):
    """Anexa uma foto ao cliente (jpeg/png/webp/gif, max 5MB)."""
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=422, detail="Tipo de arquivo nao permitido. Use JPEG, PNG, WEBP ou GIF.")

    c = await db.get(Cliente, cliente_id)
    if not c:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cliente nao encontrado")

    # Mesmo bloqueio da edicao: cliente com alteracao pendente nao aceita mudancas
    if c.status_endereco == "atualizando":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cliente com alteração pendente de aprovação. Anexe fotos após a aprovação."
        )

    data = await file.read()
    if len(data) > MAX_IMAGE_SIZE:
        raise HTTPException(status_code=422, detail="Arquivo muito grande. Maximo 5 MB.")
    if not storage.validar_imagem(data):
        raise HTTPException(status_code=422, detail="Conteudo do arquivo nao e uma imagem valida.")

    foto_id = uuid4()
    ext = storage.EXT_POR_TIPO.get(file.content_type, "jpg")
    key = f"{cliente_id}/{foto_id}.{ext}"
    url = storage.upload_file(key, data, file.content_type, bucket=storage.CLIENTES_BUCKET)

    foto = ClienteFoto(id=foto_id, cliente_id=cliente_id, url=url)
    db.add(foto)
    await db.commit()
    await db.refresh(foto)
    return FotoOut.from_orm(foto)


@router.delete("/clientes/{cliente_id}/fotos/{foto_id}", response_model=ClienteOut)
async def deletar_foto(
    cliente_id: UUID,
    foto_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("editar")),
):
    """Remove uma foto especifica do cliente (apaga do MinIO e do banco)."""
    c = await db.get(Cliente, cliente_id)
    if not c:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cliente nao encontrado")

    # Mesmo bloqueio da edicao: cliente com alteracao pendente nao aceita mudancas
    if c.status_endereco == "atualizando":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cliente com alteração pendente de aprovação. A alteração será liberada após a aprovação."
        )

    foto = await db.get(ClienteFoto, foto_id)
    if not foto or foto.cliente_id != cliente_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Foto nao encontrada")

    # Remove do MinIO
    info = storage.extract_key_from_url(foto.url)
    if info:
        storage.delete_file(info[1], bucket=info[0])

    await db.delete(foto)
    await db.commit()
    await db.refresh(c)
    return await _cliente_out(db, c)


# --- Submissoes pendentes (aprovador) ---

CAMPOS_EDITAVEIS = [
    "codigo", "nome_razao_social", "telefone", "pessoa_contato", "cep", "rua",
    "numero", "bairro", "cidade", "estado", "latitude", "longitude",
    "ponto_referencia", "observacao"
]


class AlteracaoOut(BaseModel):
    id: UUID
    cliente_id: UUID
    cliente_codigo: str | None = None
    cliente_nome: str = ""
    snapshot: dict
    # Endereco/dados atuais do cliente no momento da submissao (para comparar no diff)
    cliente_atual: dict = {}
    motorista_nome: str
    motorista_empresa: str = "AC"
    status: str
    observacao_revisao: str | None = None
    created_at: str
    revisado_at: str | None = None
    revisado_por_nome: str | None = None
    revisado_por_empresa: str | None = None

    @classmethod
    def from_orm(
        cls,
        a: ClienteAlteracao,
        cliente_atual: dict | None = None,
        cliente_codigo: str | None = None,
        cliente_nome: str = "",
    ):
        return cls(
            id=a.id,
            cliente_id=a.cliente_id,
            cliente_codigo=cliente_codigo,
            cliente_nome=cliente_nome,
            snapshot=a.snapshot or {},
            cliente_atual=cliente_atual or {},
            motorista_nome=a.motorista_nome,
            motorista_empresa=a.motorista_empresa or "AC",
            status=a.status,
            observacao_revisao=a.observacao_revisao,
            created_at=a.created_at.isoformat() if a.created_at else None,
            revisado_at=a.revisado_at.isoformat() if a.revisado_at else None,
            revisado_por_nome=a.revisado_por_nome,
            revisado_por_empresa=a.revisado_por_empresa,
        )


def _cliente_atual_dict(c: Cliente | None) -> dict:
    """Extrai os campos de endereco atuais do cliente para o diff."""
    if c is None:
        return {}
    def _f(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return v
    return {
        "rua": c.rua,
        "numero": c.numero,
        "bairro": c.bairro,
        "cep": c.cep,
        "cidade": c.cidade,
        "estado": c.estado,
        "latitude": _f(c.latitude),
        "longitude": _f(c.longitude),
        "ponto_referencia": c.ponto_referencia,
        "observacao": c.observacao,
        "telefone": c.telefone,
        "pessoa_contato": c.pessoa_contato,
        "nome_razao_social": c.nome_razao_social,
    }


async def _cliente_atual_com_enderecos(db: AsyncSession, c: Cliente | None) -> dict:
    """Dados atuais do cliente (flat + enderecos/contatos) para o diff."""
    base = _cliente_atual_dict(c)
    if c is None:
        return base
    enderecos = await _carregar_enderecos(db, c.id)
    base["enderecos"] = []
    for e in enderecos:
        contatos = (await _carregar_contatos(db, [e.id])).get(e.id, [])
        base["enderecos"].append(_endereco_para_dict(e, contatos))
    return base


def _endereco_para_dict(e: ClienteEndereco, contatos: list[ClienteContato]) -> dict:
    """Serializa um endereco (com contatos) para JSON/diff/snapshot."""
    def _f(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return v
    return {
        "nome": e.nome,
        **{k: _f(getattr(e, k)) for k in COLUNAS_ENDERECO_XLSX},
        "contatos": [{"nome": ct.nome, "telefone": ct.telefone} for ct in contatos],
    }


class RecusarBody(BaseModel):
    observacao: str | None = Field(default=None, max_length=500)


class EditarAlteracaoBody(BaseModel):
    """Edita o snapshot antes de aprovar. Campos opcionais."""
    nome_razao_social: constr(strip_whitespace=True, max_length=150) | None = None
    telefone: constr(strip_whitespace=True, max_length=20) | None = None
    pessoa_contato: constr(strip_whitespace=True, max_length=100) | None = None
    cep: constr(strip_whitespace=True, max_length=10) | None = None
    rua: constr(strip_whitespace=True, max_length=150) | None = None
    numero: constr(strip_whitespace=True, max_length=20) | None = None
    bairro: constr(strip_whitespace=True, max_length=100) | None = None
    cidade: constr(strip_whitespace=True, max_length=100) | None = None
    estado: constr(strip_whitespace=True, min_length=2, max_length=2) | None = None
    latitude: float | None = None
    longitude: float | None = None
    ponto_referencia: constr(strip_whitespace=True, max_length=200) | None = None
    observacao: constr(strip_whitespace=True, max_length=2000) | None = None


@router.get("/alteracoes", response_model=list[AlteracaoOut])
async def listar_alteracoes(
    status_filter: str | None = Query(default=None, alias="status"),
    empresa: str | None = Query(default=None, alias="empresa"),
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("aprovar")),
):
    """Lista submissoes (default: todas; passe ?status=pendente para pendentes).
    ?empresa=AC|SIN filtra pelo MOTORISTA que solicitou a alteracao."""
    stmt = select(ClienteAlteracao, Cliente).join(
        Cliente, ClienteAlteracao.cliente_id == Cliente.id, isouter=True
    ).order_by(ClienteAlteracao.created_at.desc())
    if status_filter:
        stmt = stmt.where(ClienteAlteracao.status == status_filter)
    if empresa:
        stmt = stmt.where(ClienteAlteracao.motorista_empresa == empresa)
    result = await db.execute(stmt)
    rows = result.all()
    saida = []
    for (a, c) in rows:
        cliente_atual = await _cliente_atual_com_enderecos(db, c) if c else {}
        saida.append(AlteracaoOut.from_orm(
            a,
            cliente_atual=cliente_atual,
            cliente_codigo=c.codigo if c else None,
            cliente_nome=c.nome_razao_social if c else "",
        ))
    return saida


async def _get_alteracao(db: AsyncSession, alt_id: UUID) -> ClienteAlteracao:
    a = await db.get(ClienteAlteracao, alt_id)
    if not a:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Submissao nao encontrada")
    return a


async def _aplica_snapshot(c: Cliente, snapshot: dict, db: AsyncSession):
    """Aplica o snapshot aos campos editaveis do cliente e enderecos/contatos."""
    for campo in CAMPOS_EDITAVEIS:
        if campo in snapshot:
            setattr(c, campo, snapshot[campo])

    # Conjunto de enderecos/contatos propostos (substitui o atual)
    enderecos = snapshot.get("enderecos")
    if enderecos is not None:
        enderecos_data = []
        for ed in enderecos:
            ed = dict(ed)
            await _geocode_endereco_dict(ed)
            enderecos_data.append(ed)
        await _substituir_enderecos(db, c.id, enderecos_data)
        _espelhar_primeiro_endereco(c, enderecos_data)
    else:
        # Snapshot antigo (antes do modelo de multiplos enderecos): reflete os
        # campos flat no primeiro endereco para manter consistencia.
        flat = {k: snapshot[k] for k in CAMPOS_ENDERECO_SET if k in snapshot}
        if flat:
            atual = await _carregar_enderecos(db, c.id)
            if atual:
                for k, v in flat.items():
                    setattr(atual[0], k, v)

    # Se ficou sem coords (e endereco existe) tenta geocoding
    if (c.latitude is None and c.longitude is None) and (c.rua or c.cidade or c.cep):
        coords = await geocode_endereco(
            rua=c.rua, numero=c.numero, bairro=c.bairro,
            cidade=c.cidade, estado=c.estado, cep=c.cep
        )
        if coords:
            c.latitude = coords[0]
            c.longitude = coords[1]


# --- Notificações (sino / WebSocket) ---

async def _notificar_aprovadores(
    db: AsyncSession,
    *,
    cliente: Cliente,
    alteracao_id: UUID,
    motorista_nome: str,
) -> list[tuple[User, Notificacao]]:
    """Cria Notificacao (sem commit) para os aprovadores que devem revisar.

    Aprovador = role admin (todas as empresas) OU user com permissão 'aprovar'
    da MESMA empresa do motorista que solicitou.
    Retorna lista de (user, notificacao) para o chamador enviar após o commit.
    """
    result = await db.execute(select(User).where(User.is_active == True))
    alvos: list[tuple[User, Notificacao]] = []
    empresa_motorista = cliente.alterado_por_empresa or "AC"
    for u in result.scalars().all():
        eh_admin = u.role == "admin"
        tem_perm = "aprovar" in (u.permissions or [])
        if not (eh_admin or tem_perm):
            continue
        if not eh_admin and (u.empresa or "AC") != empresa_motorista:
            continue
        n = criar_notificacao(
            db,
            u.id,
            tipo="nova_alteracao",
            titulo="Nova solicitação para aprovar",
            mensagem=(
                f"{cliente.nome_razao_social} aguarda revisão de alteração "
                f"de endereço (solicitada por {motorista_nome})."
            ),
            link="/aprovacoes?status=pendente",
            cliente_id=cliente.id,
            alteracao_id=alteracao_id,
        )
        alvos.append((u, n))
    return alvos


def _notificar_motorista(
    db: AsyncSession,
    *,
    alteracao: ClienteAlteracao,
    cliente_id: UUID,
    cliente_nome: str,
    tipo: str,
    titulo: str,
    mensagem: str,
) -> Notificacao:
    """Cria Notificacao (sem commit) para o motorista que solicitou a alteracao."""
    return criar_notificacao(
        db,
        alteracao.motorista_user_id,
        tipo=tipo,
        titulo=titulo,
        mensagem=mensagem,
        link=f"/clientes/editar?cliente={cliente_id}",
        cliente_id=cliente_id,
        alteracao_id=alteracao.id,
    )


@router.post("/alteracoes/{alt_id}/aprovar", response_model=AlteracaoOut)
async def aprovar_alteracao(
    alt_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: Any = Depends(require_permission("aprovar")),
):
    """Aprova a submissao: aplica o snapshot no cliente e fecha como 'aprovado'."""
    a = await _get_alteracao(db, alt_id)
    if a.status != "pendente":
        raise HTTPException(status_code=400, detail=f"Submissao ja {a.status}")

    c = await db.get(Cliente, a.cliente_id)
    if not c:
        raise HTTPException(status_code=404, detail="Cliente nao existe mais")

    await _aplica_snapshot(c, a.snapshot or {}, db)

    # Status volta a aprovado, mas alterado_por_* continua sendo quem PEDIU (o motorista),
    # para a pesquisa/histórico mostrar quem solicitou a mudanca.
    c.status_endereco = "aprovado"

    a.status = "aprovado"
    a.revisado_at = datetime.now(timezone.utc)
    a.revisado_por_user_id = UUID(user["user_id"])
    a.revisado_por_nome = user["name"]
    a.revisado_por_empresa = user.get("empresa") or "AC"

    # Notifica o motorista que solicitou
    notif = _notificar_motorista(
        db,
        alteracao=a,
        cliente_id=c.id,
        cliente_nome=c.nome_razao_social,
        tipo="aprovada",
        titulo="Sua alteração foi aprovada",
        mensagem=f"A alteração de endereço de '{c.nome_razao_social}' foi aprovada por {user['name']}.",
    )

    await db.commit()
    await db.refresh(a)
    await enviar_notificacao(a.motorista_user_id, notif)
    return AlteracaoOut.from_orm(
        a,
        cliente_atual=await _cliente_atual_com_enderecos(db, c),
        cliente_codigo=c.codigo,
        cliente_nome=c.nome_razao_social,
    )


@router.post("/alteracoes/{alt_id}/recusar", response_model=AlteracaoOut)
async def recusar_alteracao(
    alt_id: UUID,
    body: RecusarBody,
    db: AsyncSession = Depends(get_db),
    user: Any = Depends(require_permission("aprovar")),
):
    """Recusa a submissao: mantem o endereco atual do cliente."""
    a = await _get_alteracao(db, alt_id)
    if a.status != "pendente":
        raise HTTPException(status_code=400, detail=f"Submissao ja {a.status}")

    c = await db.get(Cliente, a.cliente_id)
    # Se cliente existe, volta o status dele para 'aprovado' (sem alterar endereco)
    if c:
        c.status_endereco = "aprovado"
        c.alterado_por_user_id = None
        c.alterado_por_nome = None
        c.alterado_por_empresa = None
        c.alterado_em = None

    a.status = "recusado"
    a.observacao_revisao = body.observacao
    a.revisado_at = datetime.now(timezone.utc)
    a.revisado_por_user_id = UUID(user["user_id"])
    a.revisado_por_nome = user["name"]
    a.revisado_por_empresa = user.get("empresa") or "AC"

    # Notifica o motorista que solicitou
    notif = _notificar_motorista(
        db,
        alteracao=a,
        cliente_id=c.id if c else a.cliente_id,
        cliente_nome=c.nome_razao_social if c else "Cliente",
        tipo="recusada",
        titulo="Sua alteração foi recusada",
        mensagem=(
            f"A alteração de endereço foi recusada por {user['name']}."
            + (f" Motivo: {body.observacao}" if body.observacao else "")
        ),
    )

    await db.commit()
    await db.refresh(a)
    await enviar_notificacao(a.motorista_user_id, notif)
    return AlteracaoOut.from_orm(
        a,
        cliente_atual=await _cliente_atual_com_enderecos(db, c) if c else {},
        cliente_codigo=c.codigo if c else None,
        cliente_nome=c.nome_razao_social if c else "",
    )


@router.put("/alteracoes/{alt_id}/editar", response_model=AlteracaoOut)
async def editar_e_aprovar_alteracao(
    alt_id: UUID,
    body: EditarAlteracaoBody,
    db: AsyncSession = Depends(get_db),
    user: Any = Depends(require_permission("aprovar")),
):
    """Aprovador ajusta o snapshot antes de aplicar, e ja aprova."""
    a = await _get_alteracao(db, alt_id)
    if a.status != "pendente":
        raise HTTPException(status_code=400, detail=f"Submissao ja {a.status}")

    c = await db.get(Cliente, a.cliente_id)
    if not c:
        raise HTTPException(status_code=404, detail="Cliente nao existe mais")

    # Aplica os overrides do aprovador por cima do snapshot
    snapshot = dict(a.snapshot or {})
    updates = body.model_dump(exclude_unset=True)
    for k, v in updates.items():
        snapshot[k] = v

    # Se o aprovador ajustou campos de endereco flat, espelha no 1o endereco
    # do snapshot para manter a consistencia entre os dois.
    enderecos_snap = snapshot.get("enderecos")
    if enderecos_snap:
        primeiros = [k for k in updates if k in COLUNAS_ENDERECO_XLSX]
        if primeiros:
            enderecos_snap[0] = dict(enderecos_snap[0])
            for k in primeiros:
                enderecos_snap[0][k] = updates[k]
            snapshot["enderecos"] = enderecos_snap

    await _aplica_snapshot(c, snapshot, db)
    c.status_endereco = "aprovado"
    # alterado_por_* mantém quem PEDIU (o motorista) — o revisor fica em revisado_por_*

    # Atualiza o snapshot da submissao com os ajustes e marca como 'editado'/aprovado
    a.snapshot = snapshot
    a.status = "editado"
    a.observacao_revisao = "Editado e aprovado pelo revisor"
    a.revisado_at = datetime.now(timezone.utc)
    a.revisado_por_user_id = UUID(user["user_id"])
    a.revisado_por_nome = user["name"]
    a.revisado_por_empresa = user.get("empresa") or "AC"

    # Notifica o motorista que solicitou
    notif = _notificar_motorista(
        db,
        alteracao=a,
        cliente_id=c.id,
        cliente_nome=c.nome_razao_social,
        tipo="editada",
        titulo="Sua alteração foi aprovada com ajustes",
        mensagem=f"A alteração de endereço de '{c.nome_razao_social}' foi aprovada com ajustes por {user['name']}.",
    )

    await db.commit()
    await db.refresh(a)
    await enviar_notificacao(a.motorista_user_id, notif)
    return AlteracaoOut.from_orm(
        a,
        cliente_atual=await _cliente_atual_com_enderecos(db, c),
        cliente_codigo=c.codigo,
        cliente_nome=c.nome_razao_social,
    )


@router.get("/clientes/{cliente_id}/alteracoes", response_model=list[AlteracaoOut])
async def historico_alteracoes_cliente(
    cliente_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: Any = Depends(require_permission("visualizar")),
):
    """Historico de submissoes de alteracao de um cliente especifico."""
    c = await db.get(Cliente, cliente_id)
    if not c:
        raise HTTPException(status_code=404, detail="Cliente nao encontrado")
    result = await db.execute(
        select(ClienteAlteracao)
        .where(ClienteAlteracao.cliente_id == cliente_id)
        .order_by(ClienteAlteracao.created_at.desc())
    )
    saida = []
    for a in result.scalars().all():
        saida.append(AlteracaoOut.from_orm(
            a,
            cliente_atual=await _cliente_atual_com_enderecos(db, c),
            cliente_codigo=c.codigo,
            cliente_nome=c.nome_razao_social,
        ))
    return saida